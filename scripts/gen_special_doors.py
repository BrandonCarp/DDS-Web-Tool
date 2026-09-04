#!/usr/bin/env python3
"""
Generate src/lib/pricing/data/special-doors.ts from a Clopay ALL tab.

The special order tab has always worked one way: the counter reads a total off
the Clopay portal and the tool applies the collection's margin. That still
works and is still the fallback. This adds a second path for models where
Clopay has issued a full size grid, so the common configurations can be picked
from dropdowns instead of typed.

Reads the SELL column only. PRICE, FUEL, TOTAL and MPQ are Clopay's working
columns; SELL is TOTAL at the model's margin and is what the tool quotes, so
these prices need no further margin applied — unlike the manual path, where the
margin is the whole point.

The stated margin (row 2, e.g. "43M") is verified against TOTAL/SELL on every
row rather than trusted. A grid issued before a margin change would otherwise
be applied at the old rate with nothing to show for it.

Prices are the door on its own, at 12" radius, extension springs, no lock —
the same baseline the residential grid uses. Track, spring and lock adders come
from ADDONS at quote time, exactly as they do for a stock door, which is why
the sheet's own footer values are not read: they already live in addons.ts and
duplicating them would let the two drift.

Usage:
  python3 scripts/gen_special_doors.py <sheet.xlsx> <tab> <model-key> <height>
  e.g. python3 scripts/gen_special_doors.py 4050-7FT.xlsx ALL 4050/4051/4053 7
"""
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "src/lib/pricing/data/special-doors.ts"
SIZE = re.compile(r"^(\d+)'(\d+)\"?\s*X\s*(\d+)'(\d+)\"?$", re.I)
STYLES = ("solid", "glass", "inserts")


# Rows the sheet gets wrong, corrected on Brandon's word.
#
# The 4050 at 7'0" tall prices flat from 6'0" to 7'10". The sheet's 6'0" and
# 6'2" rows carry the band BELOW that, and their GLASS cells were also $3.00
# above what the stated 43M margin gives — the two symptoms of one bad block.
# Applied here rather than edited into the generated file so that a regenerate
# does not quietly restore them.
#
# Keyed model -> height -> width -> style.
CORRECTIONS = {
    "4050/4051/4053": {
        "7": {
            "6": {"solid": 837.75, "glass": 968.32, "inserts": 1030.60},
            "6.2": {"solid": 837.75, "glass": 968.32, "inserts": 1030.60},
        }
    }
}


def width_order(key: str):
    """Sort by feet then inches.

    Not float(): "6.10" is 6 feet 10 inches, and float() reads it as 6.1 and
    files it below "6.2". The catalogue's decimal-looking width keys are not
    decimals.
    """
    ft, _, inch = key.partition(".")
    return (int(ft), int(inch or 0))


def width_key(feet: int, inches: int) -> str:
    """Catalogue width convention: whole feet as "8", part-foot as "7.6".

    The ALL tab writes some widths with a leading zero — 11'00", 11'02" — so
    the inches are parsed as a number rather than kept as text.
    """
    return str(feet) if inches == 0 else f"{feet}.{inches}"


def main():
    if len(sys.argv) != 5:
        sys.exit("usage: gen_special_doors.py <sheet.xlsx> <tab> <model-key> <height>")
    path, tab, model, height = Path(sys.argv[1]), sys.argv[2], sys.argv[3], sys.argv[4]
    ws = load_workbook(path, data_only=True)[tab]

    grid, margin, mismatched, wrong_height = {}, None, [], set()
    for raw in ws.iter_rows(values_only=True):
        for cell in raw:
            if isinstance(cell, str) and re.fullmatch(r"\d+M", cell.strip()):
                margin = int(cell.strip()[:-1])
        head = str(raw[0]).strip().upper().replace("  ", " ") if raw[0] else ""
        m = SIZE.match(head)
        if not m or raw[1] is None or len(raw) < 7:
            continue
        sell = raw[6]
        if not isinstance(sell, (int, float)):
            continue
        style = str(raw[1]).strip().lower()
        if style not in STYLES:
            continue
        if m.group(3) != height:
            wrong_height.add(f"{m.group(3)}'{m.group(4)}\"")
            continue
        total = raw[4]
        if isinstance(total, (int, float)) and margin:
            expected = round(float(total) / (1 - margin / 100), 2)
            if abs(expected - round(float(sell), 2)) > 0.02:
                mismatched.append(f"{head} {style}: SELL {sell}, TOTAL/{margin}M = {expected}")
        grid.setdefault(width_key(int(m.group(1)), int(m.group(2))), {})[style] = round(float(sell), 2)

    if margin is None:
        sys.exit("ABORT — no margin marker (e.g. '43M') found on the sheet")
    # A handful of rows off the stated margin are typos, and SELL is what DDS
    # quotes, so those are reported and used as written. MOST rows off means the
    # margin itself moved and the marker was not updated — that is a different
    # problem and it aborts, because applying a whole grid at the wrong margin
    # is not something to discover from a quote.
    total_rows = sum(len(v) for v in grid.values())
    if mismatched and total_rows and len(mismatched) / total_rows > 0.1:
        print(f"ABORT — {len(mismatched)} of {total_rows} rows disagree with the stated",
              f"{margin}M margin. That is not a typo; check the marker.", file=sys.stderr)
        for line in mismatched[:10]:
            print(f"  {line}", file=sys.stderr)
        sys.exit(1)
    if mismatched:
        print(f"{len(mismatched)} row(s) where SELL is not TOTAL at {margin}M.")
        print("Using SELL as written — it is the column DDS quotes — but these look")
        print("like typed-over formulas and are worth checking against Clopay:")
        for line in mismatched:
            print(f"  {line}")
        print()

    incomplete = {w: sorted(set(STYLES) - set(v)) for w, v in grid.items() if len(v) != 3}
    if incomplete:
        print(f"{len(incomplete)} width(s) missing a style — those cells fall back to the")
        print("manual total entry rather than quoting a guess:")
        for w, missing in sorted(incomplete.items(), key=lambda x: width_order(x[0])):
            print(f"  {w}: no {', '.join(missing)}")
        print()

    fixes = CORRECTIONS.get(model, {}).get(height, {})
    for width, styles in fixes.items():
        if width not in grid:
            print(f"NOTE: correction for {width} has no row on the sheet", file=sys.stderr)
            continue
        for style, value in styles.items():
            was = grid[width].get(style)
            if was is not None and abs(was - value) > 0.005:
                print(f"  corrected {width} {style}: {was} -> {value}")
            grid[width][style] = value
    if fixes:
        print()

    payload = {model: {height: {w: grid[w] for w in sorted(grid, key=width_order)}}}
    body = json.dumps(payload, indent=2)
    body = re.sub(r"^", "", body)
    OUT.write_text(
        "// AUTO-GENERATED from the Clopay size grids by scripts/gen_special_doors.py.\n"
        "// Do not edit by hand — re-run the script against a newer sheet.\n"
        f"//\n// Source: {path.name} ({tab} tab, {margin}M)\n"
        "//\n"
        "// Values are the SELL column: the door at the collection margin, already\n"
        "// applied. Nothing further is applied at quote time, unlike the manual path\n"
        "// where the counter enters a Clopay portal total and the margin does the work.\n"
        "//\n"
        "// Baseline is 12\" radius, extension springs, no lock. Track, spring and lock\n"
        "// adders come from ADDONS at quote time, the same ones a stock door uses.\n"
        "//\n"
        "// Shape: model -> height tier -> widthKey -> PriceTriple.\n"
        'import type { PriceTriple } from "../types";\n\n'
        "export const SPECIAL_DOORS: Record<string, Record<string, Record<string, Partial<PriceTriple>>>> =\n"
        f"  {body};\n",
        encoding="utf-8",
    )
    n = sum(len(v) for v in grid.values())
    print(f"wrote {OUT.relative_to(ROOT)} — {len(grid)} width(s), {n} price(s) at {margin}M")
    if wrong_height:
        print(f"skipped heights other than {height}': {', '.join(sorted(wrong_height))}")


if __name__ == "__main__":
    main()
