#!/usr/bin/env python3
"""
Generate src/lib/pricing/data/operator-sheet-prices.ts from LiftMaster price
sheets (NEW_LM_PRICING*.xlsx).

Third source document for operator prices, after the OPERATORS sheet
(descriptions, gen_operators.py) and the QuickBooks estimates (rates,
gen_operator_prices.py). Same rule as every other price file in the repo:
nobody types a rate by hand.

Takes any number of workbooks and reads every sheet in each, so a residential
tab and a commercial tab apply in one run the way gen_operator_prices.py
already takes several estimate PDFs. A later file wins on any key an earlier
one also set, and every override is printed rather than applied quietly.

COLUMNS ARE NOT FIXED. The RES tab puts COST in column C and SELL in E; the
COMM tab puts COST in B and SELL in F. Both are located from the header row
carrying the words COST and SELL rather than hard-coded, because the two tabs
already disagree and a third would too.

SELL is the counter price — it is the column that reproduces what the tool
already quotes. The GENERAL $ column is deliberately NOT written: nothing in
the app models a second price tier, and picking which customers get it would
be inventing pricing policy.

There is no subtotal on a price sheet, so the checksum gen_operator_prices.py
relies on does not exist here. What replaces it: every row must resolve to
exactly one catalogue description. A row matching two things aborts the write.
A row matching nothing is listed and skipped — that is a model DDS has not
added to the catalogue yet, not a parser failure.

Usage: python3 scripts/gen_operator_prices_sheet.py <sheet.xlsx> [sheet.xlsx ...]
"""
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
CATALOGUE = ROOT / "src/lib/pricing/data/operators.ts"
CATALOGUE_MANUAL = ROOT / "src/lib/pricing/data/operators-manual.ts"
OUT = ROOT / "src/lib/pricing/data/operator-sheet-prices.ts"

LENGTH = re.compile(r"\b(\d{1,2})\s*FT\b")
# Sprockets join on the chain size buried in the part number: the sheet's
# "71-1550B22LGH" and the catalogue's "1'' SPROCKET,  50B22" are one item.
SPROCKET_SHEET = re.compile(r"^71-15(50B\d+)([LQ])GH$", re.I)
SPROCKET_CAT = re.compile(r"SPROCKET,?\s*(50B\d+)", re.I)
# L and Q are the bore: 1" and 1-1/4". Every size on the 8-31 sheet prices the
# two identically, so this mapping changes no number today — but it would the
# moment they diverge, which operator-pricing.test.ts watches for.
BORE = {"L": "1", "Q": "1-1/4"}


def norm(text: str) -> str:
    """Comparison form. Folds hyphens and commas so the sheet's "I BEAM" and
    "TDC12S1BMC - 8FT" line up with the catalogue's "I-BEAM" and "TDC12S1BMC,".
    """
    return re.sub(r"[\s\-,]+", " ", text).strip().upper()


def price_key(desc: str) -> str:
    """The join key, byte-identical to priceKey() in operator-pricing.ts.

    Deliberately NOT norm(): norm() folds hyphens and commas for comparison and
    that folding must never reach a stored key. A key written "I BEAM" matches
    nothing the TypeScript side computes, and the tab reads "price not set" for
    every I-beam rail with no error anywhere to say why.
    """
    return re.sub(r"\s+", " ", desc).strip().upper()


def cell_text(value) -> str:
    """Model numbers arrive as text or as numbers.

    98022 is stored numerically, so a bare str() on a float cell yields
    "98022.0" and the join misses silently.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def find_columns(ws):
    for row in ws.iter_rows(values_only=True):
        vals = [str(c).strip().upper() if c is not None else "" for c in row]
        if "COST" in vals and "SELL" in vals:
            return {"cost": vals.index("COST"), "sell": vals.index("SELL")}
    return None


def read_sheet(ws):
    cols = find_columns(ws)
    if cols is None:
        return []
    sell_i = cols["sell"]
    # A separate qualifier column exists only when something sits between the
    # model and the first number. On COMM the qualifier is inside the model
    # cell ("MT5011U,  8FT") and column B is already COST.
    qual_i = 1 if cols["cost"] > 1 else None

    rows = []
    for i, raw in enumerate(ws.iter_rows(values_only=True), 1):
        model = cell_text(raw[0]) if raw else ""
        sell = raw[sell_i] if sell_i < len(raw) else None
        # "N/A" in the cost column leaves sell blank or non-numeric. Those are
        # models LiftMaster has not quoted, not rows to guess at.
        if not model or not isinstance(sell, (int, float)):
            continue
        qual = cell_text(raw[qual_i]) if qual_i is not None and qual_i < len(raw) else ""
        if qual.replace(".", "").isdigit():
            qual = ""
        rows.append({
            "row": i,
            "sheet": ws.title,
            "model": model,
            "qual": qual,
            # Excel hands back 575.95000000000005 for a cell reading 575.95.
            "price": round(float(sell), 2),
        })
    return rows


def read_catalogue():
    """Generated rows plus the hand-added ones.

    A price row must resolve against a model DDS sells that
    NEW_PARTS_LIST.xlsx has not listed yet, or adding it to operators-manual.ts
    still leaves it reading "price not set". Both files feed the same join,
    exactly as operator-catalogue.ts merges them at runtime.
    """
    generated = re.findall(r'\{"desc":\s*"([^"]+)"', CATALOGUE.read_text(encoding="utf-8"))
    manual = re.findall(
        r'^\s*desc:\s*"([^"]+)"',
        CATALOGUE_MANUAL.read_text(encoding="utf-8"),
        re.M,
    )
    return generated + manual


def sprocket_candidates(match, catalogue):
    size, suffix = match.group(1).upper(), match.group(2).upper()
    bore = BORE[suffix]
    out = []
    for desc in catalogue:
        found = SPROCKET_CAT.search(desc)
        if not found or found.group(1).upper() != size:
            continue
        # The catalogue writes the bore ahead of the word SPROCKET, and one row
        # uses a foot mark where every other uses an inch mark ("1' SPROCKET"
        # vs "1'' SPROCKET"). Compare the measurement, not the punctuation.
        head = desc.split("SPROCKET")[0].split("MODEL")[-1]
        measure = re.sub(r"[^0-9/\-]", "", head)
        if measure == bore:
            out.append(desc)
    return out


def candidates(row, catalogue):
    model = row["model"].upper()

    sprocket = SPROCKET_SHEET.match(model)
    if sprocket:
        return sprocket_candidates(sprocket, catalogue)

    blob = norm(f"{model} {row['qual']}")
    base = re.split(r"[,\-]", model)[0].strip().upper()
    want_len = LENGTH.search(blob)
    want_len = want_len.group(1) if want_len else None
    kind = next((k for k in ("I BEAM", "CHAIN", "BELT", "TROLLEY") if k in blob), None)
    # "7FT CHAIN RAIL" with nothing in the model column is the rail sold alone.
    # The catalogue also lists operator-plus-rail bundles ending the same way,
    # so the bundles are excluded or every rail row matches three items.
    rail_only = model.endswith("RAIL")

    out = []
    for desc in catalogue:
        nd = norm(desc)
        if rail_only:
            if "RAIL" not in nd or "ELECTRIC OPERATOR MODEL" in nd:
                continue
        elif not re.search(rf"\b{re.escape(base)}\b", nd):
            continue
        if want_len:
            found = LENGTH.search(nd)
            if not found or found.group(1) != want_len:
                continue
        elif LENGTH.search(nd):
            continue
        if kind and kind not in nd:
            continue
        out.append(desc)

    # TDC12S1BMC appears twice per length in the catalogue, once as STANDARD /
    # 20 CYCLES and once as EXTENDED / 30 CYCLES. The second is a catalogue
    # defect: the sheet lists TDC12X1BMC as its own SKU at its own price, and
    # the jackshaft pair (JHDC12S1BMC STANDARD / JHDC12X1BMC EXTENDED) sets the
    # S-is-standard, X-is-extended convention. An S-coded sheet row therefore
    # takes the STANDARD row and nothing else. The EXTENDED rows stay unmatched
    # on purpose — renaming a generated catalogue entry is a fix to
    # NEW_PARTS_LIST.xlsx, not something to infer here, and the two models are
    # roughly $490 apart.
    if len(out) > 1 and re.search(r"\d+S\d+", base):
        standard = [d for d in out if "STANDARD" in norm(d)]
        if len(standard) == 1:
            return standard
    return out


def main():
    if len(sys.argv) < 2:
        sys.exit("usage: gen_operator_prices_sheet.py <sheet.xlsx> [sheet.xlsx ...]")

    catalogue = read_catalogue()
    resolved, sources = {}, {}
    missing, ambiguous, overrides = [], [], []

    for arg in sys.argv[1:]:
        path = Path(arg)
        wb = load_workbook(path, data_only=True)
        for name in wb.sheetnames:
            for row in read_sheet(wb[name]):
                found = candidates(row, catalogue)
                if len(found) > 1:
                    ambiguous.append((path.name, row, found))
                    continue
                if not found:
                    missing.append((path.name, row))
                    continue
                key = price_key(found[0])
                origin = f"{path.name}:{row['sheet']}"
                if key in resolved and resolved[key] != row["price"]:
                    overrides.append((key, resolved[key], row["price"], sources[key], origin))
                resolved[key] = row["price"]
                sources[key] = origin

    if ambiguous:
        print("ABORT — rows matching more than one catalogue entry:", file=sys.stderr)
        for fname, row, found in ambiguous:
            print(f"  {fname}:{row['sheet']} row {row['row']}: "
                  f"{row['model']} {row['qual']} -> {found}", file=sys.stderr)
        sys.exit(1)

    for key, old, new, was, now in overrides:
        print(f"OVERRIDE  {old} ({was}) -> {new} ({now})  {key}")
    if overrides:
        print()

    if missing:
        print(f"{len(missing)} row(s) name a product not in the catalogue.")
        print("Add them to operators-manual.ts (or to NEW_PARTS_LIST.xlsx and re-run")
        print("gen_operators.py) before they can carry a price:")
        for fname, row in missing:
            qual = f" {row['qual']}" if row["qual"] else ""
            print(f"  {row['sheet']:5s} row {row['row']:3d}  "
                  f"{row['model']}{qual}  {row['price']:.2f}")
        print()

    inputs = ", ".join(Path(a).name for a in sys.argv[1:])
    lines = [
        "// AUTO-GENERATED from the LiftMaster price sheets by",
        "// scripts/gen_operator_prices_sheet.py. Do not edit by hand — re-run the",
        "// script against newer sheets when prices change.",
        "//",
        f"// Source: {inputs}",
        "//",
        "// Values are the SELL column. The sheets also carry COST and a GENERAL $",
        "// column; neither is written here, because nothing in the app models a",
        "// second price tier.",
        "//",
        "// Keys are the catalogue DESCRIPTION each sheet row resolved to, so this",
        "// file joins the same way operator-prices.ts does. See operator-pricing.ts",
        "// for which source wins when both carry the same item.",
        "//",
        "// Counter sell prices, tax excluded.",
        "",
        "export const SHEET_OPERATOR_PRICES: Record<string, number> = {",
    ]
    for key in sorted(resolved):
        lines.append(f'  "{key}": {resolved[key]},')
    lines.append("};")
    OUT.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"wrote {OUT.relative_to(ROOT)} — {len(resolved)} price(s)")


if __name__ == "__main__":
    main()
