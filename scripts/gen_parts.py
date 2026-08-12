#!/usr/bin/env python3
"""
Generate src/lib/pricing/data/parts.ts from NEW_PARTS_LIST.xlsx.

Sheet shape: column A is the part name (or a heading), column H is the
description the counter copies into QuickBooks, column I is the price. Headings
carry no description.

Per-foot items are the ones whose description ends in a comma — the sheet leaves
it open because the footage gets written on the end ("2\" RAW TRACK," becomes
"2\" RAW TRACK,  10FT"). Brush seal is per foot too but has no trailing comma,
so it is matched by name.

Per-foot parts bill as quantity 1 with the footage multiplied into the rate. The
per-foot figure never reaches the QuickBooks line. Vinyl is the exception and is
not generated here: it has its own covering logic in vinyl.ts, and it bills the
other way round — footage as the quantity, per-foot as the rate.

Usage: python3 scripts/gen_parts.py <NEW_PARTS_LIST.xlsx>
"""
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "src/lib/pricing/data/parts.ts"

# Headings that sit UNDER another heading rather than starting a new category.
SUBHEADINGS = {
    "EXTENSION CABLES",
    "TORSION CABLES",
    "EXTENSION SPRINGS, 7FT",
    "EXTENSION SPRINGS, 8FT",
    "TORSION SPRINGS, 7FT",
    "TORSION SPRINGS, 8FT",
    "TORSION SPRINGS, 9FT",
    "COMMERCIAL TRACKS",
    "RESIDENTIAL TRACKS",
}

# The sheet files all springs under one SPRINGS heading, but the counter looks
# for torsion or extension, never "springs" — so they are split into two
# categories here. That also makes the QuickBooks item name right, since the
# item is the category.
SPRING_SPLIT = "SPRINGS"


def spring_category(name, sub):
    """Which spring category a row belongs to, from its sub-heading or name."""
    text = f"{sub or ''} {name}".upper()
    if "TORSION" in text or " TOR " in f" {text} ":
        return "TORSION SPRINGS"
    if "EXTENSION" in text or " EXT " in f" {text} ":
        return "EXTENSION SPRINGS"
    return None

# Vinyl has its own module — the calculator takes a door size, not a footage.
SKIP_CATEGORIES = {"VINYL"}


def cell(row, i):
    v = row[i] if i < len(row) else None
    return "" if v is None else str(v).strip()


# The sheet writes a default pair onto every torsion spring
# ("... 23-1/4\" LONG [1] - RIGHT AND [1] - LEFT"). The counter sets the hands,
# so the suffix is stripped here and rebuilt from their numbers.
HANDS_RE = re.compile(r"\s*\[\d+\]\s*-\s*(?:RIGHT|LEFT)S?(?:\s+AND\s+\[\d+\]\s*-\s*(?:RIGHT|LEFT)S?)*\s*$", re.I)


def is_per_foot(name, desc):
    if desc.endswith(","):
        return True
    return "BRUSH SEAL" in name.upper()


def main():
    wb = load_workbook(sys.argv[1], read_only=True, data_only=True)
    rows = list(wb["PARTS"].iter_rows(max_col=12, values_only=True))

    cats, cur, sub = [], None, None
    for r in rows:
        name, desc, price = cell(r, 0), cell(r, 7), cell(r, 8)
        if not name or name == "PARTS":
            continue

        if not desc:  # a heading
            if name in SUBHEADINGS:
                sub = name
            else:
                cur = {"name": name, "items": []}
                sub = None
                cats.append(cur)
            continue

        if cur is None or cur["name"] in SKIP_CATEGORIES:
            continue
        try:
            p = float(price)
        except ValueError:
            continue
        # Rows priced at zero are labels the sheet uses to introduce a run of
        # items (e.g. "EXTENSION SPRINGS," above the spring sizes), not stock.
        if p <= 0:
            continue

        hands = bool(HANDS_RE.search(desc))
        if hands:
            desc = HANDS_RE.sub("", desc).rstrip()
        item = {"name": name, "desc": desc, "price": round(p, 2)}
        if hands:
            item["hands"] = True
        if sub:
            item["sub"] = sub
        if is_per_foot(name, desc):
            item["perFoot"] = True

        target = cur
        if cur["name"] == SPRING_SPLIT:
            split = spring_category(name, sub)
            if split:
                target = next((c for c in cats if c["name"] == split), None)
                if target is None:
                    target = {"name": split, "items": []}
                    cats.append(target)
        target["items"].append(item)

    cats = [c for c in cats if c["items"]]
    n = sum(len(c["items"]) for c in cats)
    pf = sum(1 for c in cats for i in c["items"] if i.get("perFoot"))

    body = []
    for c in cats:
        body.append(f'  {{\n    "name": {json.dumps(c["name"])},\n    "items": [')
        for it in c["items"]:
            body.append("      " + json.dumps(it, sort_keys=True) + ",")
        body.append("    ],\n  },")

    OUT.write_text(
        "// AUTO-GENERATED from NEW_PARTS_LIST.xlsx by scripts/gen_parts.py.\n"
        "// Do not edit by hand — re-run the script when the sheet changes.\n"
        "//\n"
        "// perFoot items bill as QUANTITY 1 with the footage multiplied into the\n"
        "// rate, so the per-foot figure never lands on the QuickBooks line.\n"
        "// Vinyl stop molding is deliberately absent: it takes a door size rather\n"
        "// than a footage and bills the other way round. See data/vinyl.ts.\n\n"
        'import { handSuffix } from "./torsion";\n\n'
        "export interface Part {\n"
        "  name: string;\n"
        "  /** Verbiage copied into the QuickBooks description column. */\n"
        "  desc: string;\n"
        "  /** Each price, or the per-foot rate when perFoot is set. */\n"
        "  price: number;\n"
        "  /** Heading this item sits under inside its category. */\n"
        "  sub?: string;\n"
        "  /** Sold by the linear foot — the counter enters how many. */\n"
        "  perFoot?: boolean;\n"
        "  /** Ordered by hand — the counter sets how many rights and lefts. */\n"
        "  hands?: boolean;\n"
        "}\n\n"
        "export interface PartCategory {\n"
        "  /** Also the QuickBooks item name for everything inside it. */\n"
        "  name: string;\n"
        "  items: Part[];\n"
        "}\n\n"
        "export const PART_CATEGORIES: PartCategory[] = [\n"
        + "\n".join(body)
        + "\n];\n\n"
        "/**\n"
        " * Description ready for QuickBooks.\n"
        " *\n"
        " * Per-foot parts get the footage written on (`2\" RAW TRACK,  10FT`).\n"
        " * Hand-ordered parts get the counts appended (`... [2] - RIGHTS AND [1] - LEFT`).\n"
        " */\n"
        "export function partDescription(\n"
        "  part: Part,\n"
        "  feet?: number,\n"
        "  right?: number,\n"
        "  left?: number,\n"
        "): string {\n"
        "  if (part.hands) {\n"
        "    const suffix = handSuffix(right ?? 0, left ?? 0);\n"
        "    return suffix ? `${part.desc} ${suffix}` : part.desc;\n"
        "  }\n"
        "  if (!part.perFoot || !feet) return part.desc;\n"
        "  const base = part.desc.replace(/,\\s*$/, \"\");\n"
        "  return `${base},  ${feet}FT`;\n"
        "}\n\n"
        "/** Springs are priced each — the pair shows up as quantity 2, not a doubled rate. */\n"
        "export function partQuantity(part: Part, right?: number, left?: number): number {\n"
        "  if (!part.hands) return 1;\n"
        "  return Math.max(0, Math.trunc(right ?? 0)) + Math.max(0, Math.trunc(left ?? 0));\n"
        "}\n\n"
        "/** Extended price: per-foot parts charge rate x footage, others charge each. */\n"
        "export function partPrice(part: Part, feet?: number): number {\n"
        "  if (!part.perFoot) return part.price;\n"
        "  const ft = Math.max(0, Math.trunc(feet ?? 0));\n"
        "  return Math.round(part.price * ft * 100) / 100;\n"
        "}\n"
    )
    print(f"{len(cats)} categories, {n} items, {pf} per-foot -> {OUT}", file=sys.stderr)


if __name__ == "__main__":
    main()
