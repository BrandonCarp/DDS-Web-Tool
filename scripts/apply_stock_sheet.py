#!/usr/bin/env python3
"""
Fold a single-model Clopay stock sheet into src/lib/pricing/data/stock-prices.ts.

The master stock book (new_pricing_2026_V2.xlsx) covers every model at every
tier and is what gen_stock_prices.py reads. Clopay also issues per-model sheets
mid-year — 4050-7FT.xlsx is one — that reprice a single model at a single
height. This applies one of those without touching any other model or tier.

Reads the SELL column only. PRICE, FUEL, TOTAL and MPQ are Clopay's own
working columns; SELL is TOTAL divided by the model's margin and is the only
one the tool quotes. The margin is stated on the sheet (row 2, e.g. "43M") and
is verified against TOTAL/SELL on every row rather than trusted, because a
sheet issued after a margin change would otherwise apply the old one silently.

Only rows the app already prices are updated. A size on the sheet that the app
does not carry is reported and skipped — adding a stocked size is a change to
the stock matrix in stock-colors.ts, not something a price sheet decides.

Usage:
  python3 scripts/apply_stock_sheet.py <sheet.xlsx> <tab> <model-key> <tier>
  e.g. python3 scripts/apply_stock_sheet.py 4050-7FT.xlsx STOCK 4050-4051-4053 7
"""
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
TARGET = ROOT / "src/lib/pricing/data/stock-prices.ts"
SIZE = re.compile(r"^(\d+)'(\d+)\"?\s*X\s*(\d+)'(\d+)\"?$", re.I)
STYLES = ("solid", "glass", "inserts")


def width_key(feet: int, inches: int) -> str:
    """Catalogue width convention: whole feet as "8", half-foot as "7.6"."""
    return str(feet) if inches == 0 else f"{feet}.{inches}"


def read_sheet(path: Path, tab: str):
    ws = load_workbook(path, data_only=True)[tab]
    rows, margin = {}, None
    for raw in ws.iter_rows(values_only=True):
        for cell in raw:
            if isinstance(cell, str) and re.fullmatch(r"\d+M", cell.strip()):
                margin = int(cell.strip()[:-1])
        head = str(raw[0]).strip().upper().replace("  ", " ") if raw[0] else ""
        m = SIZE.match(head)
        if not m or raw[1] is None or len(raw) < 7:
            continue
        total, sell = raw[4], raw[6]
        if not isinstance(sell, (int, float)):
            continue
        style = str(raw[1]).strip().lower()
        if style not in STYLES:
            continue
        key = width_key(int(m.group(1)), int(m.group(2)))
        rows.setdefault(key, {})[style] = (round(float(sell), 2), total)
    return rows, margin


def main():
    if len(sys.argv) != 5:
        sys.exit("usage: apply_stock_sheet.py <sheet.xlsx> <tab> <model-key> <tier>")
    path, tab, model, tier = Path(sys.argv[1]), sys.argv[2], sys.argv[3], sys.argv[4]
    rows, margin = read_sheet(path, tab)
    if margin is None:
        sys.exit("ABORT — no margin marker (e.g. '43M') found on the sheet")

    # Verify the stated margin actually produced the SELL column. A sheet cut
    # before a margin change would otherwise be applied at the old rate with
    # nothing to show for it.
    bad = []
    for width, styles in rows.items():
        for style, (sell, total) in styles.items():
            if not isinstance(total, (int, float)):
                continue
            expected = round(float(total) / (1 - margin / 100), 2)
            if abs(expected - sell) > 0.02:
                bad.append(f"{width} {style}: SELL {sell}, but TOTAL/{margin}M = {expected}")
    if bad:
        print(f"ABORT — SELL does not match the stated {margin}M margin:", file=sys.stderr)
        for b in bad[:10]:
            print(f"  {b}", file=sys.stderr)
        sys.exit(1)

    src = TARGET.read_text(encoding="utf-8")
    start = src.index(f'"{model}": {{')
    end = src.index("\n  },", start)
    block, changed, skipped = src[start:end], [], []

    for width in sorted(rows, key=float):
        wpat = re.compile(rf'(\n    "{re.escape(width)}": \{{)(.*?)(\n    \}})', re.S)
        wm = wpat.search(block)
        if not wm:
            skipped.append(f"width {width} (not a stocked size)")
            continue
        tpat = re.compile(rf'("{re.escape(tier)}": \{{ )([^}}]*)( \}})')
        tm = tpat.search(wm.group(2))
        if not tm:
            skipped.append(f"width {width} tier {tier} (model not stocked at this height)")
            continue
        current = {k: float(v) for k, v in re.findall(r'"(\w+)": ([\d.]+)', tm.group(2))}
        updated = dict(current)
        for style, (sell, _t) in rows[width].items():
            if style not in current:
                skipped.append(f"{width} {style} (not currently priced)")
                continue
            if abs(current[style] - sell) > 0.005:
                changed.append((width, style, current[style], sell))
            updated[style] = sell
        body = ", ".join(f'"{k}": {updated[k]}' for k in STYLES if k in updated)
        newtier = tm.group(1) + body + tm.group(3)
        block = block[: wm.start(2)] + wm.group(2).replace(tm.group(0), newtier) + block[wm.end(2):]

    src = src[:start] + block + src[end:]
    stamp = f"//   {model} tier {tier}: {path.name} ({tab} tab, {margin}M)"
    if stamp not in src:
        src = src.replace(
            "// Shape: model -> widthKey -> tier -> PriceTriple.",
            f"{stamp}\n// Shape: model -> widthKey -> tier -> PriceTriple.",
        )
    TARGET.write_text(src, encoding="utf-8")

    for width, style, old, new in changed:
        print(f"  {width:>5} {style:8} {old:9.2f} -> {new:9.2f}  {new - old:+8.2f}")
    print(f"\n{len(changed)} price(s) updated at {margin}M margin")
    if skipped:
        print(f"\n{len(skipped)} row(s) skipped:")
        for s in skipped:
            print(f"  {s}")


if __name__ == "__main__":
    main()
