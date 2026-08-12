#!/usr/bin/env python3
"""
Generate src/lib/pricing/data/operators.ts from the OPERATORS sheet.

Sheet shape differs from PARTS: column A is the model number, column B is the
description, and there is NO price column — DDS has not set operator pricing
yet. The tool copies descriptions and leaves the rate blank rather than
inventing a number.

The sheet's 15 headings are folded into two groups the counter actually thinks
in: the machine itself, and everything that hangs off it.

Usage: python3 scripts/gen_operators.py <NEW_PARTS_LIST.xlsx>
"""
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "src/lib/pricing/data/operators.ts"

# Heading -> group. Rails and sprockets are parts of a drive rather than the
# drive itself, so they sit with the accessories the counter adds on.
GROUPS = {
    "LIFTMASTER LOGIC 5": "Operators",
    "MAXUM OPERATORS": "Operators",
    "RESIDENTIAL BELT DRIVES": "Operators",
    "RESIDENTIAL CHAIN DRIVES": "Operators",
    "RESIDENTIAL SIDEMOUNT": "Operators",
    "LIGHT COMMERCIAL SIDEMOUNT": "Operators",
    "CONTROL PANELS": "Accessories",
    "KEYPADS": "Accessories",
    "LIFTMASTER ACCESSORIES": "Accessories",
    "PHOTOEYES": "Accessories",
    "REMOTES": "Accessories",
    "BELT RAILS": "Accessories",
    "CHAIN RAILS": "Accessories",
    "I BEAM RAILS": "Accessories",
    "SPROCKET": "Accessories",
}


def cell(row, i):
    v = row[i] if i < len(row) else None
    return "" if v is None else str(v).strip()


def main():
    wb = load_workbook(sys.argv[1], read_only=True, data_only=True)
    rows = list(wb["OPERATORS"].iter_rows(max_col=6, values_only=True))

    sections, cur = [], None
    for r in rows:
        name, desc = cell(r, 0), cell(r, 1)
        if not name:
            continue
        if not desc:  # a heading
            if name in GROUPS:
                cur = {"name": name, "group": GROUPS[name], "items": []}
                sections.append(cur)
            else:
                cur = None  # the sheet title row
            continue
        if cur is None:
            continue
        cur["items"].append({"name": name, "desc": desc})

    sections = [s for s in sections if s["items"]]
    n = sum(len(s["items"]) for s in sections)

    body = []
    for s in sections:
        body.append(
            f'  {{\n    "name": {json.dumps(s["name"])},\n'
            f'    "group": {json.dumps(s["group"])},\n    "items": ['
        )
        for it in s["items"]:
            body.append("      " + json.dumps(it, sort_keys=True) + ",")
        body.append("    ],\n  },")

    OUT.write_text(
        "// AUTO-GENERATED from NEW_PARTS_LIST.xlsx by scripts/gen_operators.py.\n"
        "// Do not edit by hand — re-run the script when the sheet changes.\n"
        "//\n"
        "// There are no prices here on purpose: DDS has not set operator pricing.\n"
        "// The tool copies the description and leaves the rate blank so nobody\n"
        "// quotes off a number the company has not agreed to.\n\n"
        "export interface Operator {\n"
        "  /** Model number, as it reads on the box. */\n"
        "  name: string;\n"
        "  /** Verbiage copied into the QuickBooks description column. */\n"
        "  desc: string;\n"
        "}\n\n"
        "export interface OperatorSection {\n"
        "  name: string;\n"
        '  /** "Operators" for the machine itself, "Accessories" for what hangs off it. */\n'
        "  group: string;\n"
        "  items: Operator[];\n"
        "}\n\n"
        "export const OPERATOR_SECTIONS: OperatorSection[] = [\n"
        + "\n".join(body)
        + "\n];\n\n"
        "export const OPERATOR_GROUPS = [\"Operators\", \"Accessories\"] as const;\n"
    )
    print(f"{len(sections)} sections, {n} items -> {OUT}", file=sys.stderr)


if __name__ == "__main__":
    main()
